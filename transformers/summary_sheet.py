from openpyxl import load_workbook
from openpyxl.styles.fonts import Font
import re
from services.storage import Storage

class SummarySheet(object):
    def __init__(self):
        super().__init__()

    def generate(self, results, department, filename = ''):
        store = Storage()
        template = store.find_template('{}.xlsm'.format(department.summary))
        _wb = load_workbook(template, read_only=False, keep_vba=True)
        self._wb = _wb
        response = {}
        ws = _wb['Overall Summary']
        self.department = department
        self.dpts = []
        self.sets = []
        self.max_session = 0

        for res in results:
            tcu = 0
            tqp = 0
            tco  = 0
            cgpa = 0
            for value in res['level_data'].values():
                tcu += value['tcu']
                tqp += value['tqp']
                tco += value['tco']
                self.max_session = max(value['session'], self.max_session)
            if tcu != 0:
                cgpa = tqp / tcu
            scorer = [cgpa, 4.495, 3.495, 2.395, 1.495]
            scorer.sort(key= lambda e: -e)
            d_class = ['11', '21', '22', '3rd', 'Pass'][scorer.index(cgpa)]
            res['result'] = {'cgpa': cgpa, 'tco': tco, 'tcu': tcu, 'tqp': tqp, 'class': d_class}
            sect = re.split('^[u,U](\d{4})/(\d{4})(\d{3})$', res['user']['mat_no'])
            res['set'] = int(sect[1])
            self.sets.append(res['set'])
            res['dpt'] = sect[2]
            self.dpts.append(res['dpt'])
            res['sn'] = int(sect[3])
            
        results.sort(key = lambda e: (-e['result']['cgpa'], -e['result']['tqp'], e['result']['tco']))
        row_shift = max(len(results) - 1, 0)

        table = ws.tables.get('SummaryTable')
        ref = table.ref
        totals = table.totalsRowCount
        if totals == None:
            totals = 0

        if row_shift > 0:
            ws.insert_rows(int(self._split_ref(ref)[4]) + 1 - totals, row_shift)
            table.ref = self._shift_range(ref, row_shift)

        top = int(self._split_ref(table.ref)[2]) + 1

        i = 0
        for out in results:
            ws['B' + str(i + top)] = out['user']['mat_no']
            ws['C' + str(i + top)] = out['user']['name']

            for l in range(0, self.department.levels):
                level = out['level_data'].get(l + 1)
                if level == None:
                    level = {'tcu': 0, 'tqp': 0}
                ws.cell(i + top, 4 + l * 2).value = level['tqp']
                ws.cell(i + top, 5 + l * 2).value = level['tcu']

            ws.cell(i + top, 1).value = ws.cell(top, 1).value
            for c in range(4 + (2 * self.department.levels), 10 + (2 * self.department.levels)):
                if re.match('(review|carryover)', str(ws.cell(top - 1, c).value).lower()) == None:
                    ws.cell(i + top, c).value = ws.cell(top, c).value

            for c in range(1, 12 + (2 * self.department.levels)):
                ws.cell(i + top, c).style = ws.cell(top, c).style
            ws.cell(i + top, 6 + (2 * self.department.levels)).value = out['result']['tco']
            ws.cell(i + top, 7 + (2 * self.department.levels)).value = out.get('review')
            ws.cell(i + top, 10 + (2 * self.department.levels)).value = out.get('user')['graduate'].upper()
            ws.cell(i + top, 11 + (2 * self.department.levels)).value = out.get('outstanding')
            
            i += 1
        self._create_degree_result_(results)
        try:
            _wb.save(filename)
            response.update({'status': 'success', 'message': 'Summary sheet generated successfully'})
        except Exception as e:
            response.update({'status': 'error', 'message': 'The file {}, is already opened by another program. Please, close the file and try again'.format(filename)})
        finally:
            _wb.close()
            _wb = None


    def _create_degree_result_(self, results):
        results.sort(key = lambda e: (e['set'], e['user']['name'], -self.dpts.count(e['dpt']), e['sn']))

        ws = self._wb['Degree Result']

        ws.cell(5, 3 + (2 * self.department.levels)).value = "SESSION: {}/{}".format(self.max_session -1, self.max_session)

        row_shift = max(len(results) - 1 + len(set(self.sets)), 0)

        table = ws.tables.get('DegreeTable')
        ref = table.ref
        totals = table.totalsRowCount
        if totals == None:
            totals = 0

        if row_shift > 0:
            ws.insert_rows(int(self._split_ref(ref)[4]) + 1 - totals, row_shift)
            table.ref = self._shift_range(ref, row_shift)

        top = int(self._split_ref(table.ref)[2]) + 1

        i = 0
        c_set = 0
        for out in results:
            if out['set'] != c_set:
                ws['B' + str(i + top)] = 'U{} SET'.format(out['set'])
                ws['B' + str(i + top)].font = Font(bold=True, size= 9)
                c_set = out['set']
                i += 1

            ws['B' + str(i + top)] = out['user']['mat_no']
            ws['C' + str(i + top)] = out['user']['name']
            ws.cell(i + top, 7 + (2 * self.department.levels)).value = out['result']['class']

            for l in range(0, self.department.levels):
                level = out['level_data'].get(l + 1)
                if level == None:
                    level = {'tcu': None, 'tqp': None}
                if level['tcu'] == 0:
                    level['tcu'] = None
                if level['tqp'] == 0:
                    level['tqp'] = None

                ws.cell(i + top, 4 + l * 2).value = level['tqp']
                ws.cell(i + top, 5 + l * 2).value = level['tcu']

            ws.cell(i + top, 1).value = ws.cell(top, 1).value
            for c in range(4 + (2 * self.department.levels), 7 + (2 * self.department.levels)):
                ws.cell(i + top, c).value = ws.cell(top, c).value

            for c in range(1, 8 + (2 * self.department.levels)):
                ws.cell(i + top, c).style = ws.cell(top, c).style
            i += 1


    def _shift_range(self, range, bottom, top=0):
        rng = self._split_ref(range)
        return rng[1] + str(int(rng[2]) + top) + ':' + rng[3] + str(int(rng[4]) + bottom)

    def _split_ref(self, ref):
        return re.split('([A-Z]+)(\d+):([A-Z]+)(\d+)', ref)

