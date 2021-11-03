from openpyxl import load_workbook
from openpyxl.comments import Comment
import re
import transformers.session_utils as session_utils
from utils import app_path


_sheetMap = {
    'name': 'B6', 'state': 'B7', 'mat_no': 'D6',
    'sex': 'G7', 'marital_status': 'E7', 'session': 'C9',
    'hod': 'B22', 'dept': 'A3', 'faculty': 'A2'
}

class _SharedData(object):
    hod = "22"
    def __init__(self, cache):
        super().__init__()
        self.cache = cache

class _Level(object):
    def __init__(self, level, session, wb, shared_data, department,  is_lead=False, is_tail=False):
        super().__init__()
        self.results = []
        self.total_shift = 0
        self.level = level
        self.session = session
        self.shared_data = shared_data
        self.is_lead = is_lead
        self.is_tail = is_tail
        self.department = department
        self.wb = wb
        self.sem_tcu = {}

        self.tqp = 0
        self.tcu = 0
        self.tco = 0

        ws = None
        self.tables = None

        if wb != None:
            ws = wb['L' + str(level * 100)]
            self.tables = []
            for i in range(0, department.semesters):
                self.results.append([])
                self.tables.append(ws.tables.get('S{}.{}'.format(level, i + 1)))
        self.ws = ws
        self.result_map = {}
        self.props = []
        self.reject = []
        if shared_data != None:
            self.reject = shared_data.cache['reject']

    def evaluate_results(self):
        results = []
        for i in range(len(self.results)):
            sem = self.results[i]
            tcu = self.sem_tcu.get(i+1)
            sem.sort(key = lambda i: (i['_session'], i['code']))
            for i in range(len(sem)-1, -1, -1):
                result = sem[i]
                # Current policy Flag comment them
                if tcu > self.department.max_cu:
                    self.shared_data.cache['review'] = True
                    result['comment'] += 'flag* [ Maximum credit load exceeded ]\n'
            results.extend(sem)
        return results
    
    def add_result(self, result, semester):
        if semester > 0:
            if semester > len(self.results):
                for i in range(semester - len(self.results)):
                    self.results.append([])
            
            ne_prop = re.split('(no-extra)', str(result['properties']))

            if self.sem_tcu.get(semester) == None:
               self.sem_tcu[semester] = 0
            
            if len(ne_prop) > 1 and self.props.count('no-extra{}'.format(semester)) == 0:
                self.props.append('no-extra{}'.format(semester))
                self.sem_tcu[semester] = 0
                for res in self.results[semester - 1]:
                    res['reason'] = 'No extra courses allowed in level {}, semester {}'.format(self.level * 100, semester)
                    self.result_map.pop(res['courseCode'])
                self.reject.extend(self.results[semester - 1])
                self.results[semester - 1].clear()

            if self.props.count('no-extra{}'.format(semester)) == 0 or len(ne_prop) > 1:
                self.results[semester - 1].append(result)
                self.result_map[result['courseCode']] = result
                au_prop = re.split('(add-unit)', str(result['properties']))
                if not (len(au_prop) > 1 or result.get('cu') == None):
                    self.sem_tcu[semester] += result['cu']
            else:
                result['reason'] = 'No extra courses allowed in level {}, semester {}'.format(self.level * 100, semester)
                self.reject.append(result)
    
    def commit_unknowns(self, results, wb, has_reason = False):
        b = 'Unknown Courses'
        t = 'Unknown'
        if has_reason:
            b = 'Flagged Results'
            t = "Flagged"
        if len(results) > 0 and wb[b] != None and wb[b].tables.get(t) != None:
            ws = wb[b]
            table = ws.tables.get(t)
            ws[_sheetMap['session']] = ' '
            ref = table.ref
            totals = table.totalsRowCount
            row_shift = max(len(results) - 1, 0)
            if row_shift > 0:
                ws.insert_rows(int(self._split_ref(ref)[4]) + 1 - totals, row_shift)
                table.ref = self._shift_range(ref, row_shift)
            top = int(self._split_ref(table.ref)[2]) + 1
            i = 0
            for result in results:
                ws['A' + str(i + top)] = re.split('&None&|&', '&{}&{}'.format(result.get('code'), result.get('courseCode')))[1]
                if result.get('title') == None:
                    result['title'] = '?'
                ws['B' + str(i + top)] = result.get('title')
                if result.get('cu') == None:
                    result['cu'] = '?'
                ws['C' + str(i + top)] = result.get('cu')
                ws['D' + str(i + top)] = result.get('score')
                ws['F' + str(i + top)] = str(result.get('session') - 1) + '/' + str(result.get('session'))
                ws['G' + str(i + top)] = result.get('reason')
                
                for c in range(1, 8):
                    ws.cell(i + top, c).style = ws.cell(top, c).style
                for c in range(5, 6):
                    ws.cell(i + top, c).value = ws.cell(top, c).value
                i += 1

    def commit(self):
        tqp = 0
        tcu = 0
        tco = 0

        self.tables = []
        refs = []
        total_shift = 0
        
        if self.ws != None:
            ws = self.ws
            ws[_sheetMap['session']] = str(self.session - 1) + '/' + str(self.session)
            
            for s in range(0, len(self.results)):
                self.tables.append(self.ws.tables.get('S{}.{}'.format(self.level, s + 1)))
        for t in range(len(self.results)):
            table = None
            shift = max(len(self.results[t]) - 1, 0)
            if self.ws != None:
                table = self.tables[t]
            if table != None:
                
                ref = table.ref
                totals = table.totalsRowCount
                if totals == None:
                    totals = 0
                if total_shift > 0 or shift > 0:
                    if shift > 0:
                        ws.insert_rows(int(self._split_ref(ref)[4]) + total_shift + 1 - totals, shift)
                    table.ref = self._shift_range(ref, total_shift + shift, total_shift)
                refs.append(int(self._split_ref(table.ref)[2]) + 1)
            else:
                refs.append(None)
            total_shift += shift
        
        if self.ws != None:
            if self.is_tail:
                if ws['G' + str(13 + total_shift + self.get_semesters_range())].value != None:
                    ws['G' + str(13 + total_shift + self.get_semesters_range())] = ws['G' + str(13 + total_shift + 
                            self.get_semesters_range())].value.replace(str(10 + self.get_semesters_range()),
                                    str(10 + total_shift + self.get_semesters_range()))
        
        self.total_shift = total_shift
         
        for c in range(len(self.results)):
            i = 0
            self.results[c].sort(key = lambda i: (i['_session'], i['code']))
            for result in self.results[c]:
                score = result.get('score')
                unit = result['cu']
                if self.ws != None:
                    top = refs[result['sem'] - 1]
                    ws['A' + str(i + top)] = result['code']
                    ws['B' + str(i + top)] = result['title']
                    ws['C' + str(i + top)] = unit
                    ws['D' + str(i + top)] = score
                    for c in range(3, 8):
                        ws.cell(i + top, c).style = ws.cell(top, c).style
                    for c in range(5, 8):
                        ws.cell(i + top, c).value = ws.cell(top, c).value
                    if result.get('comment') != '' and result.get('comment') != None:
                        ws.cell(i + top, 4).comment = Comment('Review:\n' + result.get('comment'), 'Auto', width=300)
                
                if unit != None and score != None:
                    scorer = [39.9999, 44.9999, 49.9999, 59.9999, 69.9999, score]
                    scorer.sort()
                    gp = scorer.index(score)
                    if result['flags'].count('carryover') > 0:
                        gp = min(3, gp)
                    if score < 40:
                        tco += unit
                    tcu += unit
                    tqp += gp * unit
                i += 1

        self.tqp = tqp
        self.tcu = tcu
        self.tco = tco

        if self.is_lead:
            self.shared_data.hod = str(12 + total_shift + self.get_semesters_range())
        
    def _shift_range(self, range, bottom, top=0):
        rng = self._split_ref(range)
        return rng[1] + str(int(rng[2]) + top) + ':' + rng[3] + str(int(rng[4]) + bottom)

    def _split_ref(self, ref):
        return re.split('([A-Z]+)(\d+):([A-Z]+)(\d+)', ref)

    def get_semesters_range(self):
        return 5 * self.department.semesters

    def finish(self):
        if not self.is_lead and self.ws != None:
            hod_cell = str(12 + self.total_shift + self.get_semesters_range())
            self.ws['B' + hod_cell] = self.ws['B' + hod_cell].value.replace(str(12 + self.get_semesters_range()), self.shared_data.hod)

class ResultFilter():
    def __init__(self, cache):
        if cache.get('sessions') == None:
            cache.update({'sessions': [], 'last_sem': 100, 'electives': {}, 'reject': [], 'review': False})
        self.hold = []
        self.cache = cache
        self.reject = cache['reject']
        self.reset = True
        self.data = {}
    def reset_filter(self):
        pass
    def evaluate_result(self, result):
        if self.reset:
            self.reset = False
            self.data = {}
            self.reset_filter()
        return self._evaluate_result(result)
    def _evaluate_result(self, result):
        return True
    def release_hold(self):
        hold = self.hold
        self.hold = []
        self.reset = True
        return hold

class HeadFilter(ResultFilter):
    def __init__(self, cache, results = []):
        super().__init__(cache)
        self.hold = results
    def reset_filter(self):
        self.cache.update({'electives': {}})

class CourseFilter(ResultFilter):
    def __init__(self, cache, courses, level_count, wb):
        super().__init__(cache)
        self.courses = courses
        self.level_count = level_count
        self.wb = wb
        self.reject = []

    def reset_filter(self):
        self.data.update({ 'sessions': {}, 'last_sem': 100})

    def _evaluate_result(self, result):
        course = self.courses.get(result['courseCode'])
        if course == None:
            self.reject.append(result)
            return False
        result.update(course)
        result.update({'_session': result['session'], 'comment': '', 'flags': []})
        if self.cache['last_sem'] == 100:
            l_session = self.data['sessions'].get(result['session'])
            sem_id = result['level'] + result['sem']
            if l_session == None or sem_id > l_session:
                self.data['sessions'][result['session']] = sem_id
            if sem_id > self.data['last_sem']:
                self.data['last_sem'] = sem_id
        self.hold.append(result)
        return False

    def release_hold(self):
        if self.cache['last_sem'] == 100:
            self.cache.update(session_utils.rectify(self.data['sessions'], self.level_count))
            if self.wb != None:
                unknown = _Level(None, None, None, None, None)
                unknown.commit_unknowns(self.reject, self.wb)
        return super().release_hold()

class CarryoverFilter(ResultFilter):
    def reset_filter(self):
        self.cache['result_map'] = {}
        self.data = self.cache['result_map']

    def _evaluate_result(self, result):
        map = self.data.get(result['courseCode'])
        if map == None or (map['score'] < 40 and result['session'] > map['session']):
            if map != None:
                result['comment'] = (map['comment'] + '[ session: ' + str(map['session'] - 1) + '/' 
                    + str(map['session']) + ', score: ' + str(map['score']) + ']\n')
                result['_session'] = result['session'] + 0.4
                result['code'] += '*'
                result['flags'].append('carryover')
                map['cu'] = None
            elif result['level'] != self.cache['sessions'].index(result['session']) + 1:
                result['_session'] = result['session'] + 0.2

            self.data[result['courseCode']] = result
        return super()._evaluate_result(result)

class RetakeFilter(ResultFilter):
    def reset_filter(self):
        if self.cache.get('result_map') == None:
            self.cache['result_map'] = {}
        self.data = self.cache.get('result_map')

    def _evaluate_result(self, result):
        map = self.data.get(result['courseCode'])
        if not (map == None or (map['score'] < 40 and result['session'] > map['session'])) and result['session'] != map['session']:
            self.data[result['courseCode']]['comment'] = (map['comment'] + 'flag* [ session: ' + str(result['session'] - 1) + '/' 
                + str(result['session']) + ', score: ' + str(result['score']) + ']\n')
            result['reason'] = 'Course has already been passed in session {}/{}'.format(map['session'] - 1, map['session'])
            self.reject.append(result)
            return False
        return super()._evaluate_result(result)


class ElectiveFilter(ResultFilter):
    def _evaluate_result(self, result):
        sem_id = result['level'] + result['sem']
        prop = re.split('(elective-pair):(\d+)', str(result['properties']))
        if len(prop) > 1:
            result['_session'] = result['session'] + 0.1
            if self.cache['electives'].get(sem_id) == None:
                self.cache['electives'][sem_id] = { 'all': [result['courseCode']]}
            else:
                self.cache['electives'][sem_id]['all'].append(result['courseCode'])
            self.hold.append(result)
            return False
        return super()._evaluate_result(result)

    def release_hold(self):
        # Current policy Flag comment them
        for i in range(len(self.hold)-1, -1, -1):
            result = self.hold[i]
            prop = re.split('(elective-pair):(\d+)', str(result['properties']))
            elective_pair = int(prop[2])
            sem_id = result['level'] + result['sem']
            cache = self.cache
            all = set(cache['electives'][sem_id]['all'])
            if len(all) > elective_pair:
                cache['review'] = True
                result['comment'] += 'flag* [ Excess electives taken{} ]\n'.format(all)
        return super().release_hold()

class LevelFilter(ResultFilter):
    def __init__(self, cache, levels, wb, department):
        super().__init__(cache)
        self.levels = levels
        self.level_data = _SharedData(self.cache)
        self.department = department
        self.wb = wb
        self.results = []

    def reset_filter(self):
        self.levels.clear()

    def _evaluate_result(self, result):
        level = self.cache['sessions'].index(result['session']) + 1
        sem = result['sem']
        if self.levels.get(level) == None:
            self.levels[level] = _Level(level, result['session'], self.wb, self.level_data, self.department, is_lead=level == 1, is_tail=level>=4)
        self.levels[level].add_result(result, sem)
        return False
    
    def release_hold(self):
        results = []
        for level in self.levels.values():
            _results = level.evaluate_results()
            results.extend(_results)
        if len(self.levels) == 0 and len(self.cache['sessions']) > 0:
            self.levels[level] = _Level(1, self.cache['sessions'][0], self.wb, self.level_data, self.department, is_lead=True, is_tail=False)
        results.sort(key = lambda i: (i['_session'], i['code']))
        self.hold = results
        self.results = results
        return super().release_hold()

class MissingFilter(ResultFilter):
    def __init__(self, cache, levels, courses):
        super().__init__(cache)
        self.levels = levels
        self.courses = courses

    def release_hold(self):
        results = {}
        courses = self.courses
        department = None
        wb = None
        level_data = None
        for level in self.levels.values():
            department = level.department
            wb = level.wb
            level_data = level.shared_data
            results.update(level.result_map)

        for key in courses.keys():
            sem_id = courses[key]['level'] + courses[key]['sem']
            elective_pair = 0
            prop = re.split('(elective-pair):(\d+)', str(courses[key]['properties']))
            if len(prop) > 1:
                elective_pair = int(prop[2])
            if (results.get(key) == None and sem_id <= self.cache['last_sem']
                and (elective_pair == 0 or self.cache['electives'].get(sem_id) == None or 
                len(set(self.cache['electives'][sem_id]['all'])) < elective_pair)
            ):
                result = {}
                result.update(courses[key])
                session = self.cache['sessions'][int(courses[key]['level']/100) - 1]
                result.update({'courseCode': key, 'cu': None, '_session': session + 0.5, 'session': session })

                level = self.cache['sessions'].index(result['session']) + 1
                sem = result['sem']
                if self.levels.get(level) == None:
                    self.levels[level] = _Level(level, result['session'], wb, level_data, department, is_lead=level == 1, is_tail=level>=4)
                self.levels[level].add_result(result, sem)
        return super().release_hold()

class StopFilter(ResultFilter):
    def __init__(self, cache, levels, wb):
        super().__init__(cache)
        self.levels = levels
        self.wb = wb

    def release_hold(self):
        for level in self.levels.values():
            level.commit()
        for level in self.levels.values():
            level.finish()
            self.levels[level.level] = {'tco': level.tco, 'tcu': level.tcu, 'tqp': level.tqp, 'session': level.session }
        if self.wb != None:
            unknown = _Level(None, None, None, None, None)
            unknown.commit_unknowns(self.reject, self.wb, has_reason = True)
        return super().release_hold()

class SpreadSheet(object):
    def evaluate(self, filters, index = 0):
        for result in filters[index].release_hold():
            for i in range(index + 1, len(filters)):
                if not filters[i].evaluate_result(result):
                    break
        if index < len(filters) - 1:
            self.evaluate(filters, index + 1)

    def __init__(self):
        super().__init__()
        self._wb = None
        self.scored_results = []
        self.invalid_results = []

    def generate(self, user, results, courses, department, filename = ''):
        user['name'] = (user['last_name'].upper() + ', ' + 
            user['first_name'].capitalize() + ' ' + user['other_names'].capitalize()).strip().rstrip(',')
        if department.faculty != None:
            user['faculty'] = department.faculty
        if department.department_long != None:
            user['dept'] = department.department_long
        if department.hod != None:
            user['hod'] = department.hod

        _sheetMap['hod'] = 'B{}'.format(12 + (5 * department.semesters))

        if filename != None and filename != '':
            _wb = load_workbook(app_path('static/excel/templates/{}.xlsx'.format(department.spreadsheet)))
            self._wb = _wb
            for key in user.keys():
                if _sheetMap.get(key) != None:
                    _wb['L100'][_sheetMap[key]] = user[key]
        
        cache = {}
        levels = {}
        course_filter = CourseFilter(cache, courses, department.levels, self._wb)
        carryover_filter = CarryoverFilter(cache)
        retake_filter = RetakeFilter(cache)
        elect_filter = ElectiveFilter(cache)
        level_filter = LevelFilter(cache, levels, self._wb, department)
        miss_filter = MissingFilter(cache, levels, courses)

        self.evaluate([
            HeadFilter(cache, results),
            course_filter, carryover_filter, level_filter,
            HeadFilter(cache),
            course_filter, carryover_filter, retake_filter, elect_filter, level_filter, miss_filter,
            StopFilter(cache, levels, self._wb)
        ])
         
        self.scored_results = results
        self.invalid_results = []
        self.invalid_results.extend(course_filter.reject)
        self.invalid_results.extend(cache['reject'])
        review = 'NO'
        if len(self.invalid_results) > 0 or cache['review']:
            review = 'YES'

        response = {'status': 'success', 'message': '', 'level_data': levels, 'user': user, 'review': review}
        print('Written {} results'.format(len(self.scored_results)))

        # step 4: remove unused sheets
        if self._wb != None:
            for sheet in _wb.worksheets:
                if sheet[_sheetMap['session']].value == None and sheet.title != 'L100':
                    _wb.remove(sheet)

            try:
                _wb.save(filename)
                response.update({'status': 'success', 'message': 'Spreadsheet generated successfully'})
            except Exception as e:
                response.update({'status': 'error', 'message': 'The file {}, is already opened by another program. Please, close the file and try again'.format(filename)})
            finally:
                _wb.close()
                _wb = None

        return response
