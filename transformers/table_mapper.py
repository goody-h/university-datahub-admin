from openpyxl import load_workbook
import uuid, re
# The excel magic is done here!

class TableMapper(object):
    def __init__(self, filename, table_anchor, row_check, column_map, header_map = []):
        super().__init__()
        self.batchId = self.getBatchId()
        self.filename = filename
        self.anchor = None
        self.topLeft = [0, 0]
        self.bottomRight = [0, 0]
        self.headerMap = []
        self.data_rows = []
        self.table_anchor = table_anchor
        self.column_map = column_map
        self.header_map = header_map
        self.row_check = row_check
    
    def __modify_row__(self, row):
        pass

    def __post_header_call__(self):
        pass
    
    def __pre_sheet_call__(self):
        pass

    def __post_sheet_call__(self):
        pass

    def __is_valid_header__(self):
        return True

    def is_Anchor(self, value):
        return re.search(self.table_anchor, str(value).lower()) != None

    def map_Header(self, value):
        for map in self.column_map:
            if re.search(map['search'], str(value).lower()) != None:
                return map['key']
        return 'annotation'

    def getBatchId(self):
        return self.rnd_id()

    def rnd_id(self):
        return uuid.uuid4().hex

    def _has_rows(self):
        return self.bottomRight[0] > self.topLeft[0]

    def _parse_header(self):
        for r in range(1, 21):
            for c in range(1, 21):
                cell = self.ws.cell(r, c).value
                if self.is_Anchor(cell):
                    self.anchor = [r, c]
                    break
                else:
                    for map in self.header_map:
                        if re.search(map['search'], str(cell).lower()) != None:
                            map['exec'](r, c)
                            break
            if self.anchor != None:
                break

    def _peek_right(self, pattern, row, column):
        for c in range(column + 1, column + 5):
            cell = self.ws.cell(row, c).value
            if re.search(pattern, str(cell).strip()) != None:
                return cell
        return None

    def _parse_sheet(self):
        self.__pre_sheet_call__()
        self._parse_header()
        self.__post_header_call__()
        if self.anchor != None and self.__is_valid_header__():
            self._go_left()
            self._go_right()
            self._go_down()
            self._map_Headers()
            if self._has_rows():
                self._get_rows()
        self.__post_sheet_call__()

    def _get_rows(self):
        for ur in range(self.topLeft[0] + 1, self.bottomRight[0] + 1):
            
            row_c = self.sanitize(self.ws.cell(ur, self.topLeft[1] + self.headerMap.index(self.row_check['key'])).value)
            if re.fullmatch(self.row_check['match'], row_c) == None:
                continue
            row = {'batchId': self.batchId }
            annotation = []
            for uc in range(self.topLeft[1], self.bottomRight[1] + 1):
                if self.headerMap[uc - self.topLeft[1]] != 'annotation':
                    row[self.headerMap[uc - self.topLeft[1]]] = self.sanitize(self.ws.cell(ur, uc).value)
                else:
                    annotation.append({
                        'key': self.sanitize(self.ws.cell(self.anchor[0], uc).value),
                        'value': self.sanitize(self.ws.cell(ur, uc).value)
                    })
            row['annotation'] = str(annotation)
            self.__modify_row__(row)
            self.data_rows.append(row)

    def sanitize(self, value):
        if value == None:
            return ''
        elif type(value) == str:
            return value.strip()
        else:
            return value

    def _map_Headers(self):
        for h in range(self.topLeft[1], self.bottomRight[1] + 1):
            self.headerMap.append(self.map_Header(self.ws.cell(self.anchor[0], h).value))
            
    def _go_left(self):
        self.topLeft[1] = self.anchor[1]
        for l in range(self.anchor[1] -1, 0, -1):
            if self.ws.cell(self.anchor[0], l).value == None:
                break
            else:
                self.topLeft[1] = l
    
    def _go_right(self):
        self.bottomRight[1] = self.anchor[1]
        while True:
            self.bottomRight[1] = self.bottomRight[1] + 1
            if self.ws.cell(self.anchor[0], self.bottomRight[1]).value == None:
                self.bottomRight[1] = self.bottomRight[1] - 1
                break
    
    def _go_down(self):
        self.topLeft[0] = self.anchor[0]
        self.bottomRight[0] = self.anchor[0]
        overflow = 0
        while True:
            self.bottomRight[0] = self.bottomRight[0] + 1
            if self.ws.cell(self.bottomRight[0], self.anchor[1]).value == None:
                if overflow >= 10:
                    self.bottomRight[0] = self.bottomRight[0] - 1
                    break
                else:
                    overflow += 1
            else:
                overflow = 0

    def get_data(self):
        self.wb = load_workbook(self.filename, data_only=True)
        for sheet in self.wb.worksheets:
            self.ws = sheet
            self.anchor = None
            self._parse_sheet()
        self.wb.close()
        self.wb = None
        return self.data_rows